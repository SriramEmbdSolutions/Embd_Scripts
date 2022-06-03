import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class JsonToExcel:

    def __init__(self, jsonfilepath, excelfilename, datanameinjson):
        self.jsonfilepath = jsonfilepath
        self.data_from_json = None
        self.excelfilename = './' + excelfilename + '.xlsx'
        self.datanameinjson = datanameinjson
        self.columns_names = []
        self.all_data_list = []

    def get_data_from_json(self):
        try:
            if os.path.exists(self.jsonfilepath):
                with open(self.jsonfilepath, 'r') as file_open_from_json:
                    self.data_from_json = json.load(file_open_from_json)
            else:
                raise Exception('Error no file found...')
        except Exception as e:
            print('Error occured, error: ', e)

    def get_columns_names_from_json_keys(self):
        if self.data_from_json != None:
            for item_row in self.data_from_json[self.datanameinjson]:
                for item in item_row:
                    if item not in self.columns_names:
                        self.columns_names.append(item)
        else:
            raise Exception("Error, data from json is empty or not initialized...")

    def get_data_rows_from_json(self):
        if self.data_from_json != None:
            for row_data in self.data_from_json[self.datanameinjson]:
                single_row_data = []
                for i in row_data.values():
                    single_row_data.append(i)
                self.all_data_list.append(single_row_data)
        else:
            raise Exception("Error, data from json is empty or not initialized...")

    def convert_json_to_excel(self):
        try:
            if os.path.exists(self.excelfilename):
                os.remove(self.excelfilename, dir_fd=None)
            else:
                work_book = Workbook()
                work_sheet = work_book.active
                work_sheet.title = 'JSON'

                self.get_data_from_json()
                self.get_columns_names_from_json_keys()
                self.get_data_rows_from_json()

                for i in range(0, len(self.columns_names)):
                    column_letter = get_column_letter(i + 1)
                    work_sheet[column_letter + "1"] = self.columns_names[i]

                for i in range(0, len(self.all_data_list)):
                    work_sheet.append(self.all_data_list[i])

                work_book.save(self.excelfilename)

        except Exception as e:
            print('Error occured. Error type: ', e)

if __name__ == '__main__':
    app = JsonToExcel('./jsonfile.json', 'newexcelfile', 'Employees')
    app.convert_json_to_excel()
    print(app.columns_names)