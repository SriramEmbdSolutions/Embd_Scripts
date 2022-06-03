import json
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class JsonToExcelClass:
    def __init__(self, jsonfilename):
        self.jsonfilename = jsonfilename
        self.data = None
        self.headers = None
        self.all_data = None
        self.dataset = None
        self.datasets = []
        self.firstcolumnname = 'id'

    def get_data_from_json_file(self):
        with open(str(self.jsonfilename), 'r+') as f:
            self.data = json.load(f)
            return self.data

    def set_first_column_header_name(self, firstcolumnname):
        self.firstcolumnname = firstcolumnname

    def get_major_data_keys_from_json_file(self):
        if self.data == None:
            self.get_data_from_json_file()
        for i in self.data.keys():
            if i not in self.datasets:
                self.datasets.append(i)
        self.dataset = self.datasets[0]
        return self.datasets

    def get_headers(self):
        self.headers = [str(self.firstcolumnname).capitalize()]
        if self.dataset == None:
            self.get_major_data_keys_from_json_file()

        for i in self.data[self.dataset]:
            if i not in self.headers:
                self.headers.append(str(i).capitalize())
        return self.headers

    def get_all_data_from_json(self):
        result = []
        for i in self.data:
            if str(i).isdigit():
                a = int(i)
            else:
                a = str(i)
            data_row_list = [a]
            for j in self.data[i].values():
                data_row_list.append(j)
            result.append(data_row_list)
        self.all_data = result
        return result

    def get_data_for_specific_id(self, id):
        result = []
        if self.data == None:
            self.get_from_json_file()
        for item in self.data[id].values():
            result.append(str(item).strip())
        return result

    def convert_json_into_excel(self, newexcelfilename):
        if self.data == None:
            self.get_data_from_json_file()
        if self.all_data == None:
            self.get_all_data_from_json()
        if self.headers == None:
            self.get_headers()
        if '/' in str(newexcelfilename):
            newexcelfilename = str(newexcelfilename).split("/")[-1]

        if "\\" in str(newexcelfilename):
            newexcelfilename = str(newexcelfilename).split("\\")[-1]

        if str(newexcelfilename).split(".")[-1] == 'xlsx':
            newexcelfilename = str(newexcelfilename).replace('.xlsx', '')

        newexcelfilename = './../../FilesToBeEdit/' + newexcelfilename + '.xlsx'

        if os.path.exists(newexcelfilename):
            os.remove(newexcelfilename, dir_fd=None)
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = 'JSON'

        for i in range(0, len(self.headers)):
            column_letter = get_column_letter(i + 1)
            work_sheet[column_letter + "1"] = self.headers[i]
            work_sheet[column_letter + "1"].font = Font(bold=True)

        for i in range(0, len(self.all_data)):
            work_sheet.append(self.all_data[i])

        work_book.save(newexcelfilename)
        return True

# if __name__ == '__main__':
#     data = get_data_from_json_file('../../FilesToBeEdit/students_excel_file.json')
#     headers = get_columns_names(data, '1', 'id')
#     all_data = get_all_data_from_json(data)
#     #print(all_data)
#     convert_json_into_excel(all_data, headers, 'C:/user/hellomilf')

# if __name__ == '__main__':
#     app = JsonToExcelClass('../../FilesToBeEdit/students_excel_file.json')
#     app.set_first_column_header_name('Number')
#     app.convert_json_into_excel('testintesting')
#     print(app.datasets)
#     print(app.firstcolumnname)

# if __name__ == '__main__':
#     app = JsonToExcelClass('./jsonfile.json')
#     app.set_first_column_header_name('Number')
#     app.convert_json_into_excel('testintesting')
#     print(app.datasets)
#    print(app.firstcolumnname)